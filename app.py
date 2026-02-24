
from flask import Flask, jsonify, render_template, request, redirect, url_for
from openpyxl import load_workbook
import os

app = Flask(__name__)
FILE = "data.xlsx"
asset_columns = [
    "AssetID","AssetName","AssetType","PurchaseDate",
    "PurchaseCost","Category","Brand","Location",
    "Status","VendorName"
]

employee_columns = [
    "EmployeeID","EmployeeName","Department",
    "Designation","BloodGroup","Email",
    "PhoneNumber","Location","Status"
]

transaction_columns = [
    "TransactionID","EmployeeID","AssetID",
    "TransactionType","IssueDate",
    "ExpectedReturnDate","ActualReturnDate",
    "Condition","ApprovedBy","Remarks"
]

def sheet_map(name):
    return {
        "assets": "Assets",
        "employees": "Employees",
        "transactions": "Transactions"
    }[name]

def get_sheet(name):
    wb = load_workbook(FILE)
    sheet = wb[name]
    data = list(sheet.values)
    headers = data[0]
    rows = data[1:]
    return wb, sheet, headers, rows

@app.route('/')
def dashboard():
    wb = load_workbook(FILE)
    counts = {
        "assets": wb["Assets"].max_row - 1,
        "employees": wb["Employees"].max_row - 1,
        "transactions": wb["Transactions"].max_row - 1
    }
    return render_template("dashboard.html", counts=counts)

@app.route('/chatbot')
def chatbot():
    return render_template("chat.html")

def generate_specific_table(headers, rows, required_columns):
    # Get column indexes
    col_indexes = [headers.index(col) for col in required_columns if col in headers]

    table_html = "<table class='table table-bordered table-striped' style='font-size:12px;'>"

    # Add headers
    table_html += "<tr>"
    for col in required_columns:
        table_html += f"<th>{col}</th>"
    table_html += "</tr>"

    # Add rows
    for row in rows:
        table_html += "<tr>"
        for index in col_indexes:
            table_html += f"<td>{row[index]}</td>"
        table_html += "</tr>"

    table_html += "</table>"

    return table_html

# ADD THIS UNIVERSAL FILTER FUNCTION 
    def filter_data(headers, rows, message):
        message = message.lower()
        filtered = rows

        for i, col in enumerate(headers):
            col_lower = col.lower()

            # BELOW condition
            if f"{col_lower} below" in message:
                try:
                    value = float(message.split(f"{col_lower} below")[-1].strip())
                    filtered = [r for r in filtered if r[i] and float(r[i]) < value]
                except:
                    return []

            # ABOVE condition
            elif f"{col_lower} above" in message:
                try:
                    value = float(message.split(f"{col_lower} above")[-1].strip())
                    filtered = [r for r in filtered if r[i] and float(r[i]) > value]
                except:
                    return []

            # EQUAL / TEXT match
            elif col_lower in message:
                value = message.split(col_lower)[-1].strip()
                filtered = [r for r in filtered if r[i] and value in str(r[i]).lower()]

        return filtered

def chatbot_logic(message):
    message = message.lower().strip()

    # -------------------------
    # LOAD SHEETS USING YOUR METHOD
    # -------------------------
    _, _, asset_headers, asset_rows = get_sheet(sheet_map("assets"))
    _, _, emp_headers, emp_rows = get_sheet(sheet_map("employees"))
    _, _, trans_headers, trans_rows = get_sheet(sheet_map("transactions"))

    # -------------------------
    # COUNT ASSETS
    # -------------------------
    if "how many assets" in message or "count assets" in message:
        return f"Total Assets: {len(asset_rows)}"

    # -------------------------
    # SHOW ALL ASSETS
    # -------------------------
    elif message == "assets" or "show assets" in message:

        table_html = "<table border='1' style='border-collapse:collapse; width:100%; font-size:12px;'>"

        table_html += "<tr>"
        for header in asset_headers:
            table_html += f"<th style='padding:6px; background:#f2f2f2;'>{header}</th>"
        table_html += "</tr>"

        for row in asset_rows[:5]:
            table_html += "<tr>"
            for cell in row:
                table_html += f"<td style='padding:6px;'>{cell}</td>"
            table_html += "</tr>"

        table_html += "</table>"

        return table_html

# -------------------------
# SEARCH ASSET
# -------------------------
    elif message.startswith("search asset"):
        keyword = message.replace("search asset", "").strip().lower()
        filtered = [r for r in asset_rows if keyword in str(r).lower()]

        if not filtered:
            return "No matching asset found."

        return generate_specific_table(asset_headers, filtered, asset_columns)
    
# REPLACE ASSET SECTION 
    elif message.startswith("show assets"):

            filtered = filter_data(asset_headers, asset_rows, message)

            print("MESSAGE:", message)
            print("FILTERED DATA:", filtered)

            if not filtered:
                return "No matching assets found."

            return generate_specific_table(asset_headers, filtered, asset_columns)
    # -------------------------
    # COUNT EMPLOYEES
    # -------------------------
    elif "how many employees" in message or "count employees" in message:
            
        return f"Total Employees: {len(emp_rows)}"

    # -------------------------
    # SHOW ALL EMPLOYEES
    # -------------------------
    elif message == "employees" or "show employees" in message:
        table_html = "<table border='1' style='border-collapse:collapse; width:100%; font-size:12px;'>"
        # add headers
        table_html += "<tr>"
        for header in emp_headers:
            table_html += f"<th style='padding:6px; background:#f2f2f2;'>{header}</th>"
        table_html += "</tr>"
        # add first 5 rows 
        for row in emp_rows[:5]:
            table_html += "<tr>"
            for cell in row:
                table_html += f"<td style='padding:6px;'>{cell}</td>"
            table_html += "</tr>"

        table_html += "</table>"

        return table_html

    # ----------------------
    # SEARCH EMPLOYEE
    # -------------------------

    elif message.startswith("search employee"):
        keyword = message.replace("search employee", "").strip().lower()
        filtered = [r for r in emp_rows if keyword in str(r).lower()]

        if not filtered:
            return "No matching employee found."

        return generate_specific_table(emp_headers, filtered, employee_columns)
    
   # REPLACE EMPLOYEE SECTION
    elif message.startswith("show employees"):

        filtered = filter_data(emp_headers, emp_rows, message)

        if not filtered:
            return "No matching employees found."

        return generate_specific_table(emp_headers, filtered, employee_columns) 

    # -------------------------
    # SEARCH EMPLOYEE BY ID
    # -------------------------
    elif "employee" in message:
        words = message.split()
        for word in words:
            for row in emp_rows:
                if str(row[0]).lower() == word.lower():  # assuming Employee ID is first column
                    return f"Employee Found:\n{row}"
        return "Employee not found."

    # -------------------------
    # SEARCH ASSET BY NAME
    # -------------------------
    elif "asset" in message:
        for row in asset_rows:
            for cell in row:
                if message in str(cell).lower():
                    return f"Asset Found:\n{row}"
        return "Asset not found."

    # -------------------------
    # SHOW TRANSACTIONS
    # -------------------------
    elif message == "transactions" or "show transactions" in message:
        table_html = "<table border='1' style='border-collapse:collapse; width:100%; font-size:12px;'>"

        # Add headers
        table_html += "<tr>"
        for header in trans_headers:
            table_html += f"<th style='padding:6px; background:#f2f2f2;'>{header}</th>"
        table_html += "</tr>"

        # Add rows (show first 5)
        for row in trans_rows[:5]:
            table_html += "<tr>"
            for cell in row:
                table_html += f"<td style='padding:6px;'>{cell}</td>"
            table_html += "</tr>"

        table_html += "</table>"

        return table_html
    # -------------------------
    # SEARCH TRANSACTION
    # -------------------------
    elif message.startswith("search transaction"):
            keyword = message.replace("search transaction", "").strip().lower()
            # filtered = [r for r in trans_rows if keyword in str(r).lower()]
            filtered = []

            for row in trans_rows:
                row_text = str(row).lower()
                print(f"Checking row: {row_text}")
                if keyword in row_text:
                    filtered.append(row)

            if not filtered:
                return "No matching transaction found."

            return generate_specific_table(trans_headers, filtered, transaction_columns)
    
   # REPLACE TRANSACTION SECTION
    elif message.startswith("show transactions"):

        filtered = filter_data(trans_headers, trans_rows, message)

        if not filtered:
            return "No matching transactions found."

        return generate_specific_table(trans_headers, filtered, transaction_columns)
    # -------------------------
    # DEFAULT RESPONSE
    # -------------------------
    else:
        return "Sorry, I didn't understand that. You can ask about assets, employees, or transactions."

@app.route("/get_response", methods=["POST"])
def get_response():
    user_message = request.json["message"]
    response = chatbot_logic(user_message)
    return jsonify({"response": response})

@app.route('/<name>')
def view(name):
    wb, sheet, headers, rows = get_sheet(sheet_map(name))

    search = request.args.get("search")
    if search:
        rows = [r for r in rows if search.lower() in str(r).lower()]

    assets = []
    employees = []
    if name == "transactions":
        _, _, _, assets = get_sheet("Assets")
        _, _, _, employees = get_sheet("Employees")

    return render_template("table.html",
                           name=name,
                           headers=headers,
                           rows=rows,
                           assets=assets,
                           employees=employees)

@app.route('/add/<name>', methods=["POST"])
def add(name):
    wb = load_workbook(FILE)
    sheet = wb[sheet_map(name)]
    sheet.append(list(request.form.values()))
    wb.save(FILE)
    return redirect(url_for('view', name=name))

@app.route('/delete/<name>/<int:row_id>')
def delete(name, row_id):
    wb = load_workbook(FILE)
    sheet = wb[sheet_map(name)]
    sheet.delete_rows(row_id + 2)
    wb.save(FILE)
    return redirect(url_for('view', name=name))

@app.route('/edit/<name>/<int:row_id>')
def edit(name, row_id):
    wb, sheet, headers, rows = get_sheet(sheet_map(name))
    return render_template("edit.html",
                           name=name,
                           headers=headers,
                           row=rows[row_id],
                           row_id=row_id)

@app.route('/update/<name>/<int:row_id>', methods=["POST"])
def update(name, row_id):
    wb = load_workbook(FILE)
    sheet = wb[sheet_map(name)]
    for col, value in enumerate(request.form.values(), start=1):
        sheet.cell(row=row_id + 2, column=col).value = value
    wb.save(FILE)
    return redirect(url_for('view', name=name))


@app.route('/chat', methods=['GET', 'POST'])
def chat():
    if request.method == "GET":
        return "Chatbot is running. Use POST request."

    message = request.form.get("message").lower()
    
    wb = load_workbook(FILE)

    if "how many assets" in message:
        sheet = wb["Assets"]
        count = sheet.max_row - 1
        return {"reply": f"Total assets are {count}"}

    return {"reply": "Sorry, I didn't understand that."}

if __name__ == "__main__":
    app.run(debug=True)
